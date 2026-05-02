import asyncio
import csv
import io
import json
import logging
from contextlib import suppress
from datetime import datetime, timezone
from typing import Any

import openpyxl
import pandas as pd
from aiokafka import AIOKafkaConsumer
from fastapi import APIRouter
from fastapi.responses import Response

from app.config import settings

logger = logging.getLogger("monolith.analytics")

router = APIRouter(prefix="/analytics", tags=["analytics"])

TOPICS = [
    "purchase.created", "purchase.voting.started", "purchase.voting.closed",
    "purchase.voting.tie", "purchase.vote.cast", "purchase.vote.changed",
    "purchase.candidate.added", "purchase.cancelled",
    "payment.topup.completed", "payment.hold.created", "payment.committed", "payment.released",
    "commission.held", "commission.committed", "commission.released",
    "escrow.created", "escrow.deposited", "escrow.confirmed", "escrow.released", "escrow.disputed",
    "review.created", "complaint.filed", "complaint.resolved", "user.auto_blocked",
    "search.query",
]

event_store: list[dict[str, Any]] = []
purchase_stats: dict[str, dict] = {}
payment_stats: dict[str, dict] = {}
commission_stats: dict[str, dict] = {}
escrow_stats: dict[str, dict] = {}
reputation_stats: dict[str, dict] = {}
search_stats: dict[str, Any] = {"total_queries": 0, "avg_latency_ms": 0}

_consumer_task: asyncio.Task | None = None


def _generate_purchases_xlsx() -> bytes:
    rows = [e for e in event_store if "purchase" in e.get("topic", "")]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Events"
    headers = ["Timestamp", "Topic", "Purchase ID", "Session ID", "Winner ID", "Total Votes", "User ID"]
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    for row in rows:
        p = row.get("payload", {})
        ws.append([
            row.get("received_at", ""), row.get("topic", ""),
            p.get("purchaseId", ""), p.get("sessionId", ""),
            p.get("winnerId", ""), p.get("totalVotes", ""),
            p.get("userId", p.get("organizerId", "")),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 50
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _generate_payments_csv() -> bytes:
    rows = [e for e in event_store if "payment" in e.get("topic", "")]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Timestamp", "Topic", "User ID", "Wallet ID", "Amount", "Currency", "Transaction ID", "Purchase ID"])
    for row in rows:
        p = row.get("payload", {})
        writer.writerow([
            row.get("received_at", ""), row.get("topic", ""),
            p.get("userId", ""), p.get("walletId", ""),
            p.get("amount", ""), p.get("currency", "RUB"),
            p.get("transactionId", ""), p.get("purchaseId", ""),
        ])
    return buf.getvalue().encode("utf-8-sig")


def _generate_votes_xlsx() -> bytes:
    votes = [
        e for e in event_store
        if e.get("topic") in ("purchase.vote.cast", "purchase.vote.changed", "purchase.voting.closed")
    ]
    df = pd.DataFrame([
        {
            "topic": e["topic"],
            "session_id": e["payload"].get("sessionId", ""),
            "purchase_id": e["payload"].get("purchaseId", ""),
            "user_id": e["payload"].get("userId", ""),
            "candidate_id": e["payload"].get("candidateId", e["payload"].get("newCandidateId", "")),
            "winner_id": e["payload"].get("winnerId", ""),
            "total_votes": e["payload"].get("totalVotes", 0),
            "ts": e["received_at"],
        }
        for e in votes
    ])
    buf = io.BytesIO()
    if not df.empty:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Votes", index=False)
            cast_df = df[df["topic"] == "purchase.vote.cast"]
            if not cast_df.empty:
                pivot = cast_df.groupby(["session_id", "candidate_id"]).size().reset_index(name="vote_count")
                pivot.to_excel(writer, sheet_name="Vote Tally", index=False)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Votes"
        ws.append(["No data yet"])
        wb.save(buf)
    return buf.getvalue()


async def _process_event(topic: str, payload: dict) -> None:
    event_store.append({"topic": topic, "payload": payload, "received_at": datetime.now(timezone.utc).isoformat()})

    if "purchaseId" in payload:
        pid = payload["purchaseId"]
        ps = purchase_stats.setdefault(pid, {"events": 0, "votes": 0, "status": "unknown"})
        ps["events"] += 1
        if topic == "purchase.vote.cast":
            ps["votes"] += 1
        if topic == "purchase.voting.closed":
            ps["winner"] = payload.get("winnerId")
            ps["total_votes"] = payload.get("totalVotes", 0)

    if "walletId" in payload or "userId" in payload:
        uid = payload.get("userId") or payload.get("walletId")
        ps2 = payment_stats.setdefault(uid, {"total_held": 0, "total_committed": 0, "total_released": 0})
        amount = payload.get("amount", 0)
        if topic == "payment.hold.created":
            ps2["total_held"] += amount
        elif topic == "payment.committed":
            ps2["total_committed"] += amount
        elif topic == "payment.released":
            ps2["total_released"] += amount

    if topic.startswith("commission."):
        pid = payload.get("purchaseId", "unknown")
        cs = commission_stats.setdefault(pid, {"held": 0, "committed": 0, "released": 0, "percent": 0})
        amount = payload.get("amount", 0)
        if topic == "commission.held":
            cs["held"] += amount
            cs["percent"] = payload.get("percent", 0)
        elif topic == "commission.committed":
            cs["committed"] += amount
        elif topic == "commission.released":
            cs["released"] += amount

    if topic.startswith("escrow."):
        pid = payload.get("purchaseId", "unknown")
        es = escrow_stats.setdefault(pid, {"total_deposited": 0, "confirmations": 0, "required": 0, "status": "active"})
        if topic == "escrow.deposited":
            es["total_deposited"] += payload.get("amount", 0)
        elif topic == "escrow.confirmed":
            es["confirmations"] = payload.get("confirmationsReceived", 0)
            es["required"] = payload.get("confirmationsRequired", 0)
        elif topic == "escrow.released":
            es["status"] = "released"
        elif topic == "escrow.disputed":
            es["status"] = "disputed"

    if topic in ("review.created", "complaint.filed", "complaint.resolved", "user.auto_blocked"):
        target_id = payload.get("targetId") or payload.get("userId", "unknown")
        rs = reputation_stats.setdefault(target_id, {"reviews": 0, "avg_rating": 0.0, "complaints": 0, "blocked": False})
        if topic == "review.created":
            rs["reviews"] += 1
            rating = payload.get("rating", 0)
            rs["avg_rating"] = ((rs["avg_rating"] * (rs["reviews"] - 1)) + rating) / rs["reviews"]
        elif topic == "complaint.filed":
            rs["complaints"] += 1
        elif topic == "user.auto_blocked":
            rs["blocked"] = True

    if topic == "search.query":
        search_stats["total_queries"] += 1


async def _consumer_loop() -> None:
    retry_delay = 5
    while True:
        consumer = AIOKafkaConsumer(
            *TOPICS,
            bootstrap_servers=settings.kafka_brokers,
            group_id="monolith-analytics-group",
            auto_offset_reset="latest",
            enable_auto_commit=True,
            value_deserializer=lambda m: json.loads(m.decode("utf-8")),
        )
        try:
            await consumer.start()
            logger.info("Analytics consumer started")
            async for msg in consumer:
                with suppress(Exception):
                    await _process_event(msg.topic, msg.value)
        except asyncio.CancelledError:
            break
        except Exception as exc:
            logger.error("Analytics consumer error: %s — retrying in %ds", exc, retry_delay)
            await asyncio.sleep(retry_delay)
            retry_delay = min(retry_delay * 2, 60)
        finally:
            with suppress(Exception):
                await consumer.stop()


def start_analytics_consumer() -> asyncio.Task:
    global _consumer_task
    _consumer_task = asyncio.create_task(_consumer_loop())
    return _consumer_task


async def stop_analytics_consumer() -> None:
    global _consumer_task
    if _consumer_task:
        _consumer_task.cancel()
        with suppress(asyncio.CancelledError):
            await _consumer_task
        _consumer_task = None


@router.get("/stats/purchases", summary="Get purchase event statistics")
async def get_purchase_stats():
    return {"success": True, "data": purchase_stats}


@router.get("/stats/payments", summary="Get payment statistics")
async def get_payment_stats():
    return {"success": True, "data": payment_stats}


@router.get("/stats/commissions", summary="Get commission statistics")
async def get_commission_stats():
    return {"success": True, "data": commission_stats}


@router.get("/stats/escrow", summary="Get escrow statistics")
async def get_escrow_stats():
    return {"success": True, "data": escrow_stats}


@router.get("/stats/reputation", summary="Get reputation statistics")
async def get_reputation_stats():
    return {"success": True, "data": reputation_stats}


@router.get("/stats/search", summary="Get search query statistics")
async def get_search_stats():
    return {"success": True, "data": search_stats}


@router.get("/stats/summary", summary="Get overall analytics summary")
async def get_summary():
    return {
        "success": True,
        "data": {
            "total_events": len(event_store),
            "purchases_tracked": len(purchase_stats),
            "users_tracked": len(payment_stats),
            "commissions_tracked": len(commission_stats),
            "escrow_accounts_tracked": len(escrow_stats),
            "reputation_profiles_tracked": len(reputation_stats),
            "search_queries": search_stats["total_queries"],
        },
    }


@router.post("/reports/generate", summary="Trigger report generation")
async def trigger_report_generation():
    """Manually trigger generation of all analytics reports."""
    loop = asyncio.get_event_loop()
    xlsx = await loop.run_in_executor(None, _generate_purchases_xlsx)
    csv_data = await loop.run_in_executor(None, _generate_payments_csv)
    return {
        "success": True,
        "message": "Reports generated",
        "sizes": {"purchases_xlsx": len(xlsx), "payments_csv": len(csv_data)},
    }


@router.get("/reports/purchases/download", summary="Download purchases report as XLSX")
async def download_purchases_xlsx():
    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, _generate_purchases_xlsx)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchases.xlsx"},
    )


@router.get("/reports/payments/download", summary="Download payments report as CSV")
async def download_payments_csv():
    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, _generate_payments_csv)
    return Response(
        content=data,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=payments.csv"},
    )


@router.get("/reports/votes/download", summary="Download voting summary as XLSX")
async def download_votes_xlsx():
    loop = asyncio.get_event_loop()
    data = await loop.run_in_executor(None, _generate_votes_xlsx)
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=votes.xlsx"},
    )
